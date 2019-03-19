import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.get_sheet_by_name('Sheet')

update_list = {'Celery': 1.19,'Garlic': 3.07, 'Lemon': 1.27}

for row_num in range(2, sheet.max_row + 1):
    produce_name = sheet.cell(row=row_num, column=1).value

    if produce_name in update_list:
        sheet.cell(row=row_num, column=2).value = update_list[produce_name]

wb.save('out/produceSales_update.xlsx')
